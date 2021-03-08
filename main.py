import requests

from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

output_filename = 'pandora_export.xlsx'
url = 'http://pandora.hundredsoul.com'
user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.192 Safari/537.36'
cookies = {'lang': 'kr'}

headers = {
    'user-agent': user_agent
}

write_wb = Workbook()


def get_items(header, name, no):
    item_session = requests.session()
    item_session.headers = headers
    item_session.cookies.update(cookies)

    with item_session:
        item_response = item_session.get('{}/list.php?id={}'.format(url, no))
        item_html = item_response.text
        item_soup = BeautifulSoup(item_html, 'html.parser')

        # Create Header
        item_ws = write_wb.create_sheet('{} - {}'.format(header, name))
        item_ws.append(['불변의 유황', '무한의 수은', '영원의 소금', '시도', '제작된 수', '(%)'])

        # Find a pagination
        max_page = 0
        pages = item_soup.findAll('ul', {'class': 'pagination'})
        for page in pages:
            page_nums = page.findAll('li')
            for page_num in page_nums:
                page_num_href = page_num.find('a')
                page_num_text = page_num_href.getText()
                if '...' in page_num_text:
                    max_page = str(page_num_href['href'])
                    max_page = int(max_page.split(sep='&')[3].replace('p=', ''))

        # All page data
        for i in range(1, max_page):
            page_response = item_session.get('{}/list.php?id={}&p={}'.format(url, no, i))
            page_html = page_response.text
            page_soup = BeautifulSoup(page_html, 'html.parser')

            # Find a data tables
            items = []
            item_tables = page_soup.findAll('table')
            for item_table in item_tables:
                item_tbody = item_table.find('tbody')
                for item_tr in item_tbody.findAll('tr'):
                    raw_data = []
                    for item_td in item_tr.findAll('td'):
                        raw = str(item_td.getText()).strip()
                        raw_data.append(raw)
                    items.append(raw_data)

            for _item in items:
                item_ws.append(_item)


if __name__ == '__main__':
    session = requests.session()
    session.headers = headers
    session.cookies.update(cookies)

    with session:
        response = session.get(url)
        html = response.text
        soup = BeautifulSoup(html, 'html.parser')

        item_types = soup.findAll('div', {'class': 'col-lg-12'})

        for item in item_types:
            item_type = item.findAll('h1')
            type_header = str()

            for _header in item_type:
                type_header = _header.getText()
                type_header = str(type_header).strip()
                type_header = type_header.replace('[', '(')
                type_header = type_header.replace(']', ')')

            if len(type_header) > 0:
                type_numbers = item.findAll('a')
                for type_number in type_numbers:
                    item_number = str(type_number['href']).strip()
                    item_number = item_number.replace('javascript:eq(\'', '')
                    item_number = item_number.replace('\');', '')
                    item_name = str(type_number.getText()).strip()
                    item_name = item_name.replace('\n', '')
                    item_name = item_name.replace('                                                        ', ' ')
                    item_name = item_name.replace('[', '(')
                    item_name = item_name.replace(']', ')')

                    get_items(type_header, item_name, item_number)

                write_wb.save(output_filename)

                wb = load_workbook(output_filename)
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
                wb.save(output_filename)
